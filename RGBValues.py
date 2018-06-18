# -*- coding: utf-8 -*-
"""
Spyder Editor

This is a temporary script file.
"""

import os
from PIL import Image
from openpyxl.chart import (
    ScatterChart,
    Reference,
    Series,

)
from openpyxl import *





experiment_date = input('What is the experiment date? Write in MM-DD-YY format.')
Main_Directory = 'C:\\Users\\Undergrunt\\Box\\Josh data\\PT RGB\\'
os.chdir(Main_Directory)

image_list = os.listdir(experiment_date)



workbook = load_workbook('C:\\Users\\Undergrunt\\Box\\Josh data\\Purple Trad RGB Graphs.xlsx')
ws = workbook.create_sheet(experiment_date)
ws.cell(row = 1, column = 1).value = 'Purple Trad Experiments'
ws.cell(row = 3, column = 1).value = 'Date'
ws.cell(row = 4, column = 1).value = experiment_date
ws.cell(row = 5, column = 1).value = 'Pixel Bin'

for j in range(256):
    ws.cell(row = j + 6, column = 1).value = j

i = 0
k = 0

dic = {}
number_files = len(image_list)

for image in image_list:
    image_name = image_list[i]
    im = Image.open(Main_Directory + experiment_date + '\\' + image_name,'r')
    red, green, blue = im.split()
    red_counts = red.histogram()
    green_counts = green.histogram()
    blue_counts = blue.histogram()
    
    for j in range(256):
        ws.cell(row = 4, column = 2 + k).value = image_name[:-4]
        ws.cell(row = 4, column = 3 + k).value = image_name[:-4]
        ws.cell(row = 4, column = 4 + k).value = image_name[:-4]
        ws.cell(row = 5, column = 2 + k).value = 'Red'
        ws.cell(row = j + 6, column = 2 + k).value = red_counts[j]
        ws.cell(row = 5, column = 3 + k).value = 'Green'
        ws.cell(row = j + 6, column = 3 + k).value = green_counts[j]
        ws.cell(row = 5, column = 4 + k).value = 'Blue'
        ws.cell(row = j + 6, column = 4 + k).value = blue_counts[j]
        
    dic['Red' + '_' + image_name[:-4]] = 2 + k
    dic['Green' + '_' + image_name[:-4]] = 3 + k
    dic['Blue' + '_' + image_name[:-4]] = 4 + k
    
    k+=3
    i+=1



red_column_gcell_list = []
red_column_ecell_list = []
green_column_gcell_list = []
green_column_ecell_list = []
blue_column_gcell_list = []
blue_column_ecell_list = []

red_column_startend_gcell_list = []
green_column_startend_gcell_list = []
blue_column_startend_gcell_list = []
red_column_startend_ecell_list = []
green_column_startend_ecell_list = []
blue_column_startend_ecell_list = []


m=0
color_list = ['red','green','blue']

for element in dic:
    if element.startswith('R') and element.endswith('G') and 'Start' not in element and 'End' not in element:
        red_column_gcell_list.append(dic[element])
        chart = ScatterChart()
        chart.title = 'Guard Cell, Red Pixels'
        chart.style = 13
        chart.x_axis.title = 'Pixel Bin'
        chart.y_axis.title = 'Intensity'
        xvalues = Reference(ws, min_col=1, min_row=5, max_row=261)
        for i in red_column_gcell_list:
            values = Reference(ws, min_col=i, min_row=4, max_row=261)
            series = Series(values, xvalues, title_from_data=True)
            chart.series.append(series)
            lineProp = drawing.line.LineProperties(solidFill = drawing.colors.ColorChoice(prstClr = color_list[m]))
            series.graphicalProperties.line = lineProp
            m+=1
        ws.add_chart(chart, "A10")
        m=0
    elif element.startswith('R') and element.endswith('E') and 'Start' not in element and 'End' not in element:
        red_column_ecell_list.append(dic[element])
        chart = ScatterChart()
        chart.title = 'Epidermal Cell, Red Pixels'
        chart.style = 13
        chart.x_axis.title = 'Pixel Bin'
        chart.y_axis.title = 'Intensity'
        xvalues = Reference(ws, min_col=1, min_row=4, max_row=261)
        for i in red_column_ecell_list:
            values = Reference(ws, min_col=i, min_row=4, max_row=261)
            series = Series(values, xvalues, title_from_data=True)
            chart.series.append(series)
            lineProp = drawing.line.LineProperties(solidFill = drawing.colors.ColorChoice(prstClr = color_list[m]))
            series.graphicalProperties.line = lineProp
            m+=1
        ws.add_chart(chart, "A24")
        m=0
    elif element.startswith('G') and element.endswith('E') and 'Start' not in element and 'End' not in element:
        green_column_ecell_list.append(dic[element])
        chart = ScatterChart()
        chart.title = 'Epidermal Cell, Green Pixels'
        chart.style = 13
        chart.x_axis.title = 'Pixel Bin'
        chart.y_axis.title = 'Intensity'
        xvalues = Reference(ws, min_col=1, min_row=5, max_row=261)
        for i in green_column_ecell_list:
            values = Reference(ws, min_col=i, min_row=4, max_row=261)
            series = Series(values, xvalues, title_from_data=True)
            chart.series.append(series)
            lineProp = drawing.line.LineProperties(solidFill = drawing.colors.ColorChoice(prstClr = color_list[m]))
            series.graphicalProperties.line = lineProp
            m+=1
        ws.add_chart(chart, "J24")
        m=0
    elif element.startswith('G') and element.endswith('G') and 'Start' not in element and 'End' not in element:
        green_column_gcell_list.append(dic[element])
        chart = ScatterChart()
        chart.title = 'Guard Cell, Green Pixels'
        chart.style = 13
        chart.x_axis.title = 'Pixel Bin'
        chart.y_axis.title = 'Intensity'
        xvalues = Reference(ws, min_col=1, min_row=5, max_row=261)
        for i in green_column_gcell_list:
            values = Reference(ws, min_col=i, min_row=4, max_row=261)
            series = Series(values, xvalues, title_from_data=True)
            chart.series.append(series)
            lineProp = drawing.line.LineProperties(solidFill = drawing.colors.ColorChoice(prstClr = color_list[m]))
            series.graphicalProperties.line = lineProp
            m+=1
        ws.add_chart(chart, "J10")
        m=0
    elif element.startswith('B') and element.endswith('E') and 'Start' not in element and 'End' not in element:
        blue_column_ecell_list.append(dic[element])
        chart = ScatterChart()
        chart.title = 'Epidermal Cell, Blue Pixels'
        chart.style = 13
        chart.x_axis.title = 'Pixel Bin'
        chart.y_axis.title = 'Intensity'
        xvalues = Reference(ws, min_col=1, min_row=5, max_row=261)
        for i in blue_column_ecell_list:
            values = Reference(ws, min_col=i, min_row=4, max_row=261)
            series = Series(values, xvalues, title_from_data=True)
            chart.series.append(series)
            lineProp = drawing.line.LineProperties(solidFill = drawing.colors.ColorChoice(prstClr = color_list[m]))
            series.graphicalProperties.line = lineProp
            m+=1
        ws.add_chart(chart, "S24")
        m=0
    elif element.startswith('B') and element.endswith('G') and 'Start' not in element and 'End' not in element:
        blue_column_gcell_list.append(dic[element])
        chart = ScatterChart()
        chart.title = 'Guard Cell, Blue Pixels'
        chart.style = 13
        chart.x_axis.title = 'Pixel Bin'
        chart.y_axis.title = 'Intensity'
        xvalues = Reference(ws, min_col=1, min_row=5, max_row=261)
        for i in blue_column_gcell_list:
            values = Reference(ws, min_col=i, min_row=4, max_row=261)
            series = Series(values, xvalues, title_from_data=True)
            chart.series.append(series)
            lineProp = drawing.line.LineProperties(solidFill = drawing.colors.ColorChoice(prstClr = color_list[m]))
            series.graphicalProperties.line = lineProp
            m+=1
        ws.add_chart(chart, "S10")
        m=0
    if element == 'Red_Start_G' or element == 'Red_End_G':
        red_column_startend_gcell_list.append(dic[element])
        chart = ScatterChart()
        chart.title = 'Guard Cell, Red Pixels'
        chart.style = 13
        chart.x_axis.title = 'Pixel Bin'
        chart.y_axis.title = 'Intensity'
        xvalues = Reference(ws, min_col=1, min_row=5, max_row=261)
        for i in red_column_startend_gcell_list:
            values = Reference(ws, min_col=i, min_row=4, max_row=261)
            series = Series(values, xvalues, title_from_data=True)
            chart.series.append(series)
            lineProp = drawing.line.LineProperties(solidFill = drawing.colors.ColorChoice(prstClr = color_list[m]))
            series.graphicalProperties.line = lineProp
            m+=1
        ws.add_chart(chart, "A38")
        m=0
    elif element == 'Red_Start_E' or element == 'Red_End_E':
        red_column_startend_ecell_list.append(dic[element])
        chart = ScatterChart()
        chart.title = 'Epidermal Cell, Red Pixels'
        chart.style = 13
        chart.x_axis.title = 'Pixel Bin'
        chart.y_axis.title = 'Intensity'
        xvalues = Reference(ws, min_col=1, min_row=4, max_row=261)
        for i in red_column_startend_ecell_list:
            values = Reference(ws, min_col=i, min_row=4, max_row=261)
            series = Series(values, xvalues, title_from_data=True)
            chart.series.append(series)
            lineProp = drawing.line.LineProperties(solidFill = drawing.colors.ColorChoice(prstClr = color_list[m]))
            series.graphicalProperties.line = lineProp
            m+=1
        ws.add_chart(chart, "A52")
        m=0
    elif element == 'Green_Start_E' or element == 'Green_End_E':
        green_column_startend_ecell_list.append(dic[element])
        chart = ScatterChart()
        chart.title = 'Epidermal Cell, Green Pixels'
        chart.style = 13
        chart.x_axis.title = 'Pixel Bin'
        chart.y_axis.title = 'Intensity'
        xvalues = Reference(ws, min_col=1, min_row=5, max_row=261)
        for i in green_column_startend_ecell_list:
            values = Reference(ws, min_col=i, min_row=4, max_row=261)
            series = Series(values, xvalues, title_from_data=True)
            chart.series.append(series)
            lineProp = drawing.line.LineProperties(solidFill = drawing.colors.ColorChoice(prstClr = color_list[m]))
            series.graphicalProperties.line = lineProp
            m+=1
        ws.add_chart(chart, "J52")
        m=0
    elif element == 'Green_Start_G' or element == 'Green_End_G':
        green_column_startend_gcell_list.append(dic[element])
        chart = ScatterChart()
        chart.title = 'Guard Cell, Green Pixels'
        chart.style = 13
        chart.x_axis.title = 'Pixel Bin'
        chart.y_axis.title = 'Intensity'
        xvalues = Reference(ws, min_col=1, min_row=5, max_row=261)
        for i in green_column_startend_gcell_list:
            values = Reference(ws, min_col=i, min_row=4, max_row=261)
            series = Series(values, xvalues, title_from_data=True)
            chart.series.append(series)
            lineProp = drawing.line.LineProperties(solidFill = drawing.colors.ColorChoice(prstClr = color_list[m]))
            series.graphicalProperties.line = lineProp
            m+=1
        ws.add_chart(chart, "J38")
        m=0
    elif element == 'Blue_Start_E' or element == 'Blue_End_E':
        blue_column_startend_ecell_list.append(dic[element])
        chart = ScatterChart()
        chart.title = 'Epidermal Cell, Blue Pixels'
        chart.style = 13
        chart.x_axis.title = 'Pixel Bin'
        chart.y_axis.title = 'Intensity'
        xvalues = Reference(ws, min_col=1, min_row=5, max_row=261)
        for i in blue_column_startend_ecell_list:
            values = Reference(ws, min_col=i, min_row=4, max_row=261)
            series = Series(values, xvalues, title_from_data=True)
            chart.series.append(series)
            lineProp = drawing.line.LineProperties(solidFill = drawing.colors.ColorChoice(prstClr = color_list[m]))
            series.graphicalProperties.line = lineProp
            m+=1
        ws.add_chart(chart, "S52")
        m=0
    elif element == 'Blue_Start_G' or element == 'Blue_End_G':
        blue_column_startend_gcell_list.append(dic[element])
        chart = ScatterChart()
        chart.title = 'Guard Cell, Blue Pixels'
        chart.style = 13
        chart.x_axis.title = 'Pixel Bin'
        chart.y_axis.title = 'Intensity'
        xvalues = Reference(ws, min_col=1, min_row=5, max_row=261)
        for i in blue_column_startend_gcell_list:
            values = Reference(ws, min_col=i, min_row=4, max_row=261)
            series = Series(values, xvalues, title_from_data=True)
            chart.series.append(series)
            lineProp = drawing.line.LineProperties(solidFill = drawing.colors.ColorChoice(prstClr = color_list[m]))
            series.graphicalProperties.line = lineProp
            m+=1
        ws.add_chart(chart, "S38")
        m=0


workbook.save(filename = 'C:\\Users\\Undergrunt\\Box\\Josh data\\Purple Trad RGB Graphs.xlsx')